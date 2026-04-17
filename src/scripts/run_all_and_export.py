#!/usr/bin/env python3
"""
Script to run all three simulation methods and save results to an Excel file.
"""

import os
import sys
import subprocess
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def run_simulation(method_name, make_target, scenario, link_bps, buffer_bytes, trace_dir):
    """Run a single simulation and capture output."""
    print(f"\n{'='*60}")
    print(f"Running {method_name}...")
    print(f"{'='*60}")
    
    env = os.environ.copy()
    env['PYTHONPATH'] = '.'
    env['SCENARIO'] = scenario
    env['TRACE_DIR'] = trace_dir
    env['LINK_BPS'] = str(link_bps)
    env['BUFFER_BYTES'] = str(buffer_bytes)
    
    try:
        result = subprocess.run(
            ['make', make_target, f'SCENARIO={scenario}', f'LINK_BPS={link_bps}', f'BUFFER_BYTES={buffer_bytes}', f'TRACE_DIR={trace_dir}'],
            cwd='.',
            capture_output=True,
            text=True,
            timeout=300
        )
        
        return {
            'method': method_name,
            'status': 'Success' if result.returncode == 0 else 'Failed',
            'return_code': result.returncode,
            'stdout': result.stdout[-500:] if result.stdout else 'No output',
            'stderr': result.stderr[-500:] if result.stderr else 'No errors'
        }
    except subprocess.TimeoutExpired:
        return {
            'method': method_name,
            'status': 'Timeout',
            'return_code': -1,
            'stdout': 'Process timed out',
            'stderr': 'Process exceeded 300 second timeout'
        }
    except Exception as e:
        return {
            'method': method_name,
            'status': 'Error',
            'return_code': -1,
            'stdout': '',
            'stderr': str(e)
        }


def save_to_excel(results, output_path):
    """Save results to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Add header
    headers = ['Method', 'Status', 'Return Code', 'Last Output', 'Errors']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1).value = result['method']
        ws.cell(row=row, column=2).value = result['status']
        ws.cell(row=row, column=3).value = result['return_code']
        ws.cell(row=row, column=4).value = result['stdout']
        ws.cell(row=row, column=5).value = result['stderr']
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    
    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")


def main():
    """Main function."""
    scenario = os.environ.get('SCENARIO', 'normal_traffic')
    link_bps = int(os.environ.get('LINK_BPS', '40000000000'))
    trace_dir = os.environ.get('TRACE_DIR', 'data/traces')

    # Use per-scenario buffer default unless explicitly overridden.
    if 'BUFFER_BYTES' in os.environ:
        buffer_bytes = int(os.environ['BUFFER_BYTES'])
    else:
        sys.path.insert(0, '.')
        from traffic.config import ScenarioName, get_buffer_bytes
        try:
            buffer_bytes = get_buffer_bytes(ScenarioName(scenario))
        except (KeyError, ValueError):
            buffer_bytes = 262144
    output_file = os.environ.get('OUTPUT_FILE', 'results.xlsx')
    
    # Create output directory if it doesn't exist
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    
    methods = [
        ('CPU FIFO', 'run-cpu-fifo'),
        ('CPU Priority Queue', 'run-cpu-priority-queue'),
        ('GPU Priority Queue', 'run-gpu-priority-queue'),
    ]
    
    results = []
    for method_name, make_target in methods:
        result = run_simulation(method_name, make_target, scenario, link_bps, buffer_bytes, trace_dir)
        results.append(result)
        print(f"{method_name}: {result['status']}")
    
    # Save results to Excel
    save_to_excel(results, output_file)
    print(f"\n{'='*60}")
    print("All simulations completed!")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
